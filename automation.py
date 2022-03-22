from tkinter import E
import openpyxl as xl
wb=xl.load_workbook('sheet_openpy.xlsx')
sheet=wb['Sheet1']

for i in range(2,sheet.max_row+1):
    sheet.cell(i,1).value=(sheet.cell(i,1).value)*0.9
    

sheet.delete_cols(5)
wb.save('sheet_openpy.xlsx')
